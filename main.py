from database import connect_db
from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import shutil
import os
from starlette.middleware.sessions import SessionMiddleware
import bcrypt

SECRET_KEY = os.getenv("SECRET_KEY")

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

app.add_middleware(
    SessionMiddleware,
    secret_key=SECRET_KEY,
    max_age=1800
)

@app.get("/", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="login.html",
        context={}
    )


failed_attempts = {}

@app.post("/login")
def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...)
):
    # Account lock after 5 wrong attempts
    if failed_attempts.get(username, 0) >= 5:
        return {"error": "Account locked. Try later."}

    conn = connect_db()

    if conn is None:
        return {"error": "Database connection failed"}

    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT * FROM users
        WHERE username=%s
    """, (username,))

    user = cursor.fetchone()
    conn.close()

    if user and bcrypt.checkpw(
        password.encode("utf-8"),
        user["password"].encode("utf-8")
    ):
        # Reset failed attempts after successful login
        failed_attempts[username] = 0

        request.session["user"] = user["username"]
        request.session["role"] = user["role"]

        if user["role"] == "admin":
            return RedirectResponse(
                url="/admin_dashboard",
                status_code=303
            )

        elif user["role"] == "teacher":
            return RedirectResponse(
                url="/teacher_dashboard",
                status_code=303
            )

        elif user["role"] == "parent":
            return RedirectResponse(
                url=f"/parent_dashboard/{user['student_id']}",
                status_code=303
            )

    # Increase failed attempts
    failed_attempts[username] = failed_attempts.get(username, 0) + 1

    return {"status": "Login Failed"}

@app.get("/students")
def get_students():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT student_name, class_name, mobile, yearly_fee
        FROM students
    """)

    students = cursor.fetchall()

    conn.close()

    return students


# ADD attendance FIRST
@app.post("/add-attendance")
def add_attendance(student_id: int, status: str):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO attendance (student_id, date, status)
        VALUES (%s, CURDATE(), %s)
    """, (student_id, status))

    conn.commit()
    conn.close()

    return {
        "status": "success",
        "message": "Attendance Added"
    }


# GET attendance AFTER add route
@app.get("/attendance/{student_id}")
def get_attendance(student_id: int):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT * FROM attendance
        WHERE student_id=%s
    """, (student_id,))

    records = cursor.fetchall()

    conn.close()

    return records


@app.get("/fees/{student_id}")
def get_fees(student_id: int):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT * FROM fees
        WHERE student_id=%s
    """, (student_id,))

    fee = cursor.fetchone()

    conn.close()

    return fee

@app.post("/add-student")
def add_student(
    admission_no: str,
    student_name: str,
    class_name: str,
    father_name: str,
    mobile: str,
    yearly_fee: float,
    discount: float
):
    conn = connect_db()
    cursor = conn.cursor()

    final_fee = yearly_fee - discount

    cursor.execute("""
        INSERT INTO students
        (admission_no, student_name, class_name, father_name, mobile, yearly_fee, discount, final_fee)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        admission_no,
        student_name,
        class_name,
        father_name,
        mobile,
        yearly_fee,
        discount,
        final_fee
    ))

    conn.commit()
    conn.close()

    return {
        "status": "success",
        "message": "Student Added"
    }

@app.post("/add-marks")
def add_marks(
    student_id: int,
    subject: str,
    marks: float,
    total_marks: float
):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO exams
        (student_id, subject, marks, total_marks)
        VALUES (%s,%s,%s,%s)
    """, (
        student_id,
        subject,
        marks,
        total_marks
    ))

    conn.commit()
    conn.close()

    return {
        "status": "success",
        "message": "Marks Added"
    }

@app.get("/marks/{student_id}")
def get_marks(student_id: int):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT subject, marks, total_marks
        FROM exams
        WHERE student_id=%s
    """, (student_id,))

    result = cursor.fetchall()

    conn.close()

    return result

@app.delete("/delete_student/{student_id}")
def delete_student(student_id: int):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM students WHERE id=%s", (student_id,))
    conn.commit()

    conn.close()

    return {"status": "success"}

@app.post("/update_student")
def update_student(
    id: int,
    name: str,
    class_name: str,
    mobile: str,
    fees: str
):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE students
        SET student_name=%s,
            class_name=%s,
            mobile=%s,
            yearly_fee=%s
        WHERE id=%s
    """, (name, class_name, mobile, fees, id))

    conn.commit()
    conn.close()

    return RedirectResponse(
        url="/web_students",
        status_code=303
    )

@app.post("/add_fee")
def add_fee(
    student_id: int,
    amount: float,
    paid_date: str,
    status: str
):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO fees (student_id, amount, paid_date, status)
        VALUES (%s,%s,%s,%s)
    """, (
        student_id,
        amount,
        paid_date,
        status
    ))

    conn.commit()
    conn.close()

    return {"status": "success"}

@app.post("/add_result")
def add_result(
    student_id: int,
    subject: str,
    marks: int,
    total_marks: int
):
    percentage = (marks / total_marks) * 100

    if percentage >= 35:
        result_status = "Pass"
    else:
        result_status = "Fail"

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO results
        (student_id, subject, marks, total_marks, percentage, result_status)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (
        student_id,
        subject,
        marks,
        total_marks,
        percentage,
        result_status
    ))

    conn.commit()
    conn.close()

    return {
        "status": "success",
        "percentage": percentage,
        "result_status": result_status

    }

@app.get("/result_list")
def result_list():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT s.student_name,
               r.subject,
               r.marks,
               r.total_marks,
               r.percentage,
               r.result_status
        FROM results r
        JOIN students s ON s.id = r.student_id
    """)

    result = cursor.fetchall()
    conn.close()

    return result

@app.get("/search_student")
def search_student(keyword: str):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT * FROM students
        WHERE student_name LIKE %s
        OR mobile LIKE %s
        OR admission_no LIKE %s
    """, (
        f"%{keyword}%",
        f"%{keyword}%",
        f"%{keyword}%"
    ))

    result = cursor.fetchall()
    conn.close()

    return result

@app.get("/parent_report/{student_id}")
def parent_report(student_id: int):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT 
            s.student_name,
            s.class_name,
            s.mobile,
            f.total_fee,
            f.pending_fee,
            a.status AS attendance_status,
            r.percentage,
            r.result_status
        FROM students s
        LEFT JOIN fees f ON s.id = f.student_id
        LEFT JOIN attendance a ON s.id = a.student_id
        LEFT JOIN results r ON s.id = r.student_id
        WHERE s.id = %s
        LIMIT 1
    """, (student_id,))

    result = cursor.fetchone()
    conn.close()

    return result

@app.get("/dashboard_stats")
def dashboard_stats():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    # Total Students
    cursor.execute("SELECT COUNT(*) as total_students FROM students")
    total_students = cursor.fetchone()["total_students"]

    # Total Fees Collected
    cursor.execute("SELECT SUM(amount) AS total_fees FROM fees")
    total_fees = cursor.fetchone()["total_fees"] or 0

    # Pending Fees
    cursor.execute("SELECT SUM(pending_fee) as pending_fees FROM fees")
    pending_fees = cursor.fetchone()["pending_fees"] or 0

    # Today Attendance
    cursor.execute("""
        SELECT COUNT(*) as today_attendance
        FROM attendance
        WHERE date = CURDATE()
    """)
    today_attendance = cursor.fetchone()["today_attendance"]

    conn.close()

    return {
        "total_students": total_students,
        "total_fees": total_fees,
        "pending_fees": pending_fees,
        "today_attendance": today_attendance
    }

@app.post("/add_teacher")
def add_teacher(
    teacher_name: str,
    mobile: str,
    subject: str,
    salary: float
):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO teachers
        (teacher_name, mobile, subject, salary)
        VALUES (%s, %s, %s, %s)
    """, (
        teacher_name,
        mobile,
        subject,
        salary
    ))

    conn.commit()
    conn.close()

    return {"status": "success"}

@app.get("/teacher_list")
def teacher_list():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM teachers")

    result = cursor.fetchall()
    conn.close()

    return result

@app.post("/delete_teacher")
def delete_teacher(id: int):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM teachers WHERE id=%s",
        (id,)
    )

    conn.commit()
    conn.close()

    return {"status": "success"}

@app.get("/admin_dashboard", response_class=HTMLResponse)
def admin_dashboard(request: Request):

    if "user" not in request.session:
        return RedirectResponse("/", status_code=303)

    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    # Students Count
    cursor.execute("SELECT COUNT(*) AS total_students FROM students")
    student_data = cursor.fetchone()
    students = student_data["total_students"]

    # Teachers Count
    cursor.execute("SELECT COUNT(*) AS total_teachers FROM teachers")
    teacher_data = cursor.fetchone()
    teachers = teacher_data["total_teachers"]

    # Fees Collected
    cursor.execute("SELECT SUM(amount) AS total_fees FROM fees")
    fee_data = cursor.fetchone()
    fees = fee_data["total_fees"] if fee_data["total_fees"] else 0

    # Results Count
    cursor.execute("SELECT COUNT(*) AS total_results FROM results")
    result_data = cursor.fetchone()
    results = result_data["total_results"]

    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="admin_dashboard.html",
        context={
            "students": students,
            "teachers": teachers,
            "fees": fees,
            "results": results
        }
    )

@app.get("/web_students", response_class=HTMLResponse)
def web_students(request: Request):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM students")
    students = cursor.fetchall()

    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="students.html",
        context={
            "students": students
        }
    )

@app.get("/add_student_web", response_class=HTMLResponse)
def add_student_web(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="add_student.html",
        context={}
    )


@app.post("/save_student_web")
async def save_student_web(
    request: Request,
    photo: UploadFile = File(...)
):
    try:
        form = await request.form()
        print("Form Data:", form)

        admission_no = form.get("admission_no")
        student_name = form.get("student_name")
        student_class = form.get("class_name")
        father_name = form.get("father_name")
        mobile = form.get("mobile")
        yearly_fee = form.get("yearly_fee")
        discount = form.get("discount") or 0

        final_fee = float(yearly_fee) - float(discount)

        # File upload security
        if not photo.filename:
            return {"error": "No file selected"}

        allowed_extensions = ["jpg", "jpeg", "png"]
        file_ext = photo.filename.split(".")[-1].lower()

        if file_ext not in allowed_extensions:
            return {"error": "Only JPG, JPEG, PNG files allowed"}

        photo_path = f"static/uploads/{photo.filename}"
        print("Saving photo:", photo_path)

        with open(photo_path, "wb") as buffer:
            shutil.copyfileobj(photo.file, buffer)

        conn = connect_db()

        if conn is None:
            return {"error": "Database connection failed"}

        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO students
            (
                admission_no,
                student_name,
                class_name,
                father_name,
                mobile,
                yearly_fee,
                discount,
                final_fee,
                photo
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            admission_no,
            student_name,
            student_class,
            father_name,
            mobile,
            yearly_fee,
            discount,
            final_fee,
            photo_path
        ))

        conn.commit()
        conn.close()

        return RedirectResponse(
            "/web_students",
            status_code=303
        )

    except Exception as e:
        print("Student Save Error:", e)
        return {"error": str(e)}

@app.get("/delete_student_web/{student_id}")
def delete_student_web(student_id: int):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM students WHERE id=%s",
        (student_id,)
    )

    conn.commit()
    conn.close()

    return RedirectResponse(
        url="/web_students",
        status_code=303
    )

@app.get("/edit_student_web/{student_id}", response_class=HTMLResponse)
def edit_student_web(
    request: Request,
    student_id: int
):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(
        "SELECT * FROM students WHERE id=%s",
        (student_id,)
    )

    student = cursor.fetchone()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="edit_student.html",
        context={
            "student": student
        }
    )

    
@app.post("/update_student_web/{student_id}")
def update_student_web(
    student_id: int,
    student_name: str = Form(...),
    class_name: str = Form(...),
    mobile: str = Form(...),
    yearly_fee: float = Form(...)
):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE students
        SET student_name=%s,
            class_name=%s,
            mobile=%s,
            yearly_fee=%s
        WHERE id=%s
    """, (
        student_name,
        class_name,
        mobile,
        yearly_fee,
        student_id
    ))

    conn.commit()
    conn.close()

    return RedirectResponse(
        url="/web_students",
        status_code=303
    )

@app.get("/web_teachers", response_class=HTMLResponse)
def web_teachers(request: Request):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM teachers")
    teachers = cursor.fetchall()

    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="teachers.html",
        context={
            "teachers": teachers
        }
    )

@app.get("/add_teacher_web", response_class=HTMLResponse)
def add_teacher_web(
    request: Request,
    success: str = None
):
    return templates.TemplateResponse(
        request=request,
        name="add_teacher.html",
        context={
            "success": success
        }
    )

@app.post("/save_teacher_web")
def save_teacher_web(
    teacher_name: str = Form(...),
    mobile: str = Form(...),
    subject: str = Form(...),
    salary: float = Form(...)
):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO teachers
        (teacher_name, mobile, subject, salary)
        VALUES (%s,%s,%s,%s)
    """, (
        teacher_name,
        mobile,
        subject,
        salary
    ))

    conn.commit()
    conn.close()

    return RedirectResponse(
        url="/add_teacher_web?success=1",
        status_code=303
    )

@app.get("/delete_teacher_web/{teacher_id}")
def delete_teacher_web(teacher_id: int):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM teachers WHERE id=%s",
        (teacher_id,)
    )

    conn.commit()
    conn.close()

    return RedirectResponse(
        url="/web_teachers",
        status_code=303
   )

@app.get("/edit_teacher_web/{teacher_id}", response_class=HTMLResponse)
def edit_teacher_web(
    request: Request,
    teacher_id: int
):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(
        "SELECT * FROM teachers WHERE id=%s",
        (teacher_id,)
    )

    teacher = cursor.fetchone()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="edit_teacher.html",
        context={
            "teacher": teacher
        }
    )

@app.post("/update_teacher_web/{teacher_id}")
def update_teacher_web(
    teacher_id: int,
    teacher_name: str = Form(...),
    mobile: str = Form(...),
    subject: str = Form(...),
    salary: float = Form(...)
):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE teachers
        SET teacher_name=%s,
            mobile=%s,
            subject=%s,
            salary=%s
        WHERE id=%s
    """, (
        teacher_name,
        mobile,
        subject,
        salary,
        teacher_id
    ))

    conn.commit()
    conn.close()

    return RedirectResponse(
        url="/web_teachers",
        status_code=303
    )

@app.get("/web_fees", response_class=HTMLResponse)
def web_fees(request: Request):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT f.*, s.student_name
        FROM fees f
        JOIN students s ON f.student_id = s.id
    """)

    fees = cursor.fetchall()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="fees.html",
        context={
            "fees": fees
        }
    )

@app.get("/add_fee_web", response_class=HTMLResponse)
def add_fee_web(
    request: Request,
    success: str = None
):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM students")
    students = cursor.fetchall()

    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="add_fee.html",
        context={
            "students": students,
            "success": success,
        }
    )

@app.post("/save_fee_web")
def save_fee_web(
    student_id: int = Form(...),
    amount: float = Form(...),
    paid_date: str = Form(...),
    status: str = Form(...)
):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    # Student total fee fetch
    cursor.execute("""
        SELECT yearly_fee
        FROM students
        WHERE id=%s
    """, (student_id,))

    student = cursor.fetchone()

    total_fee = float(student["yearly_fee"])
    paid_fee = amount
    pending_fee = total_fee - paid_fee

    cursor.execute("""
        INSERT INTO fees
        (
            student_id,
            amount,
            paid_date,
            status,
            total_fee,
            paid_fee,
            pending_fee
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s)
    """, (
        student_id,
        amount,
        paid_date,
        status,
        total_fee,
        paid_fee,
        pending_fee
    ))

    conn.commit()
    conn.close()

    return RedirectResponse(
        url="/add_fee_web?success=1",
        status_code=303
    )

@app.get("/web_pending_fees", response_class=HTMLResponse)
def web_pending_fees(request: Request):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT 
            s.id,
            s.student_name,
            s.yearly_fee,
            IFNULL(SUM(f.amount), 0) AS paid_amount
        FROM students s
        LEFT JOIN fees f ON s.id = f.student_id
        GROUP BY s.id
    """)

    fees = cursor.fetchall()

    for fee in fees:
        fee["pending_fee"] = fee["yearly_fee"] - fee["paid_amount"]

    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="pending_fees.html",
        context={"fees": fees}
    )

@app.get("/web_attendance", response_class=HTMLResponse)
def web_attendance(
    request: Request,
    success: str = None
):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM students")
    students = cursor.fetchall()

    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="attendance.html",
        context={
            "students": students,
            "success": success
        }
    )

@app.post("/save_attendance_web")
def save_attendance_web(
    student_id: int = Form(...),
    status: str = Form(...)
):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO attendance
        (student_id, date, status)
        VALUES (%s, CURDATE(), %s)
    """, (
        student_id,
        status
    ))

    conn.commit()
    conn.close()

    return RedirectResponse(
        url="/web_attendance?success=1",
        status_code=303
    )

@app.get("/web_attendance_list", response_class=HTMLResponse)
def web_attendance_list(request: Request):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT s.student_name,
               a.status,
               a.date
        FROM attendance a
        JOIN students s ON a.student_id = s.id
    """)

    attendance_list = cursor.fetchall()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="attendance_list.html",
        context={
            "attendance_list": attendance_list
        }
    )

@app.get("/web_result", response_class=HTMLResponse)
def web_result(
    request: Request,
    success: str = None
):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT id, student_name FROM students")
    students = cursor.fetchall()

    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="result.html",
        context={
            "students": students,
            "success": success
        }
    )

@app.post("/save_result_web")
def save_result_web(
    student_id: int = Form(...),
    exam_name: str = Form(...),
    subject: str = Form(...),
    marks: float = Form(...),
    total_marks: float = Form(...)
):
    percentage = (marks / total_marks) * 100

    # Result status
    result_status = "Pass"
    if percentage < 35:
        result_status = "Fail"

    # Grade calculate
    if percentage >= 75:
        grade = "A"
    elif percentage >= 60:
        grade = "B"
    elif percentage >= 35:
        grade = "C"
    else:
        grade = "F"

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO results
        (
            student_id,
            exam_name,
            subject,
            marks,
            total_marks,
            grade,
            percentage,
            result_status
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        student_id,
        exam_name,
        subject,
        marks,
        total_marks,
        grade,
        percentage,
        result_status
    ))

    conn.commit()
    conn.close()

    return RedirectResponse(
        url="/web_result?success=1",
        status_code=303
    )

@app.get("/web_result_list", response_class=HTMLResponse)
def web_result_list(request: Request):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT r.student_id,
               s.student_name,
               r.subject,
               r.marks,
               r.total_marks,
               r.percentage,
               r.result_status
        FROM results r
        JOIN students s ON r.student_id = s.id
    """)

    results = cursor.fetchall()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="result_list.html",
        context={
            "results": results
        }
    )

@app.get("/parent_details")
def parent_details():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT * FROM students
        LIMIT 1
    """)

    student = cursor.fetchone()
    conn.close()

    return student

@app.get("/parent_fee_status")
def parent_fee_status():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT amount,
               paid_date,
               status
        FROM fees
        LIMIT 1
    """)

    fee = cursor.fetchone()
    conn.close()

    return fee

@app.get("/parent_attendance")
def parent_attendance():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT status, date
        FROM attendance
        LIMIT 5
    """)

    attendance = cursor.fetchall()
    conn.close()

    return attendance

@app.get("/parent_result")
def parent_result():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT subject,
               marks,
               total_marks,
               percentage,
               result_status
        FROM results
        LIMIT 5
    """)

    results = cursor.fetchall()
    conn.close()

    return results

@app.get("/fee_receipt/{student_id}")
def fee_receipt(student_id: int):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT s.id,
               s.student_name,
               s.class_name,
               s.mobile,
               s.yearly_fee,
               f.amount
        FROM fees f
        JOIN students s ON f.student_id = s.id
        WHERE s.id=%s
    """, (student_id,))

    fee = cursor.fetchone()

    if not fee:
        return {"error": "Fee not found"}

    conn.close()

    pending_fee = fee["yearly_fee"] - fee["amount"]

    file_name = f"receipt_{student_id}.pdf"
    c = canvas.Canvas(file_name)

    c.drawString(100, 800, "SCHOOL FEE RECEIPT")
    c.drawString(100, 760, f"Student Name: {fee['student_name']}")
    c.drawString(100, 730, f"Class: {fee['class_name']}")
    c.drawString(100, 700, f"Mobile: {fee['mobile']}")
    c.drawString(100, 670, f"Total Fee: {fee['yearly_fee']}")
    c.drawString(100, 640, f"Paid Fee: {fee['amount']}")
    c.drawString(100, 610, f"Pending Fee: {pending_fee}")

    c.save()

    return FileResponse(file_name)

@app.get("/marksheet/{student_id}")
def marksheet(student_id: int):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT s.id,
               s.student_name,
               s.class_name,
               s.mobile,
               r.subject,
               r.marks,
               r.total_marks,
               r.result_status
        FROM results r
        JOIN students s ON r.student_id = s.id
        WHERE s.id=%s
    """, (student_id,))

    results = cursor.fetchall()

    if not results:
        return {"error": "No results found"}

    conn.close()

    file_name = f"marksheet_{student_id}.pdf"
    c = canvas.Canvas(file_name)

    # Outer Border
    c.rect(40, 50, 520, 760)

    # Header
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(300, 790, "REPORT CARD")

    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(300, 770, "Academic Session : 2025-26")

    # Student Info
    c.setFont("Helvetica-Bold", 12)

    c.drawString(60, 720, "Admission No.")
    c.drawString(180, 720, f": {results[0]['id']}")

    c.drawString(60, 695, "Student Name")
    c.drawString(180, 695, f": {results[0]['student_name']}")

    c.drawString(60, 670, "Class")
    c.drawString(180, 670, f": {results[0]['class_name']}")

    c.drawString(330, 720, "Roll No.")
    c.drawString(430, 720, f": {results[0]['id']}")

    c.drawString(330, 695, "Mobile")
    c.drawString(430, 695, f": {results[0]['mobile']}")

    c.drawString(330, 670, "Result Date")
    c.drawString(430, 670, ": __________")

    # Marks Table
    c.rect(60, 430, 460, 190)

    c.line(200, 430, 200, 620)
    c.line(320, 430, 320, 620)
    c.line(420, 430, 420, 620)

    c.line(60, 590, 520, 590)

    c.setFont("Helvetica-Bold", 12)

    c.drawString(80, 600, "Subject")
    c.drawString(240, 600, "Marks")
    c.drawString(350, 600, "Total")
    c.drawString(450, 600, "Status")

    # Subject rows
    y = 560
    total_obtained = 0
    grand_total = 0

    c.setFont("Helvetica", 12)

    for row in results:
        c.drawString(80, y, row['subject'])
        c.drawString(240, y, str(int(row['marks'])))
        c.drawString(350, y, str(int(row['total_marks'])))
        c.drawString(450, y, row['result_status'])

        total_obtained += row['marks']
        grand_total += row['total_marks']

        y -= 30

    # Summary
    final_percentage = (total_obtained / grand_total) * 100

    c.setFont("Helvetica-Bold", 12)

    c.drawString(
        80,
        380,
        f"Total Marks : {int(total_obtained)}/{int(grand_total)}"
    )

    c.drawString(
        80,
        350,
        f"Percentage : {round(final_percentage,2)}%"
    )

    if final_percentage >= 35:
        final_result = "Pass"
    else:
        final_result = "Fail"

    c.drawString(
        300,
        350,
        f"Result : {final_result}"
    )

    # Footer
    c.drawString(
        80,
        200,
        "Class Teacher Remark : __________"
    )

    c.drawString(
        380,
        120,
        "Principal Signature"
    )

    c.save()

    return FileResponse(file_name)

@app.get("/id_card/{student_id}")
def id_card(student_id: int):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id,
               student_name,
               class_name,
               mobile,
               photo
        FROM students
        WHERE id=%s
    """, (student_id,))

    student = cursor.fetchone()
    conn.close()

    file_name = f"id_card_{student_id}.pdf"
    c = canvas.Canvas(file_name)

    # Blue Header
    c.setFillColorRGB(0.1, 0.3, 0.9)
    c.rect(40, 700, 520, 100, fill=1)

    # School Name
    c.setFillColorRGB(1, 1, 0)
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(
        300,
        775,
        "RISING STAR ENGLISH MEDIUM SCHOOL"
    )

    # Address line 1
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica", 13)
    c.drawCentredString(
        300,
        752,
        "Nagar Jamkhed Road, Kolhewadi Phata"
    )

    # Address line 2
    c.drawCentredString(
        300,
        738,
        "Takali Kazi, Tal Dist Ahilyanagar - 414201"
    )

    # ID Title
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(300, 670, "IDENTITY CARD")

    # Photo Box
    c.rect(70, 450, 120, 160)

    if student["photo"]:
        c.drawImage(
            student["photo"],
            70,
            450,
            width=120,
            height=160
        )

    # Student Details
    c.setFont("Helvetica-Bold", 14)

    c.drawString(
        220,
        590,
        f"Student Name : {student['student_name']}"
    )

    c.drawString(
        220,
        550,
        f"Class Standard : {student['class_name']}"
    )

    c.drawString(
        220,
        510,
        f"Roll Number : {student['id']}"
    )

    c.drawString(
        220,
        470,
        f"Contact Number : {student['mobile']}"
    )

    # Footer Blue Box
    c.setFillColorRGB(0.1, 0.3, 0.9)
    c.rect(40, 100, 520, 120, fill=1)

    c.setFillColorRGB(1, 1, 1)

    c.setFont("Helvetica-Bold", 12)
    c.drawString(70, 175, "School Address")

    c.setFont("Helvetica", 10)
    c.drawString(
        70,
        150,
        "Nagar Jamkhed Road, Kolhewadi Phata"
    )

    c.drawString(
        70,
        132,
        "Takali Kazi, Tal Dist Ahilyanagar - 414201"
    )

    c.setFont("Helvetica-Bold", 12)
    c.drawString(420, 140, "Principal")

    c.save()

    return FileResponse(file_name)

@app.get("/export_fees_excel")
def export_fees_excel():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM fees")
    fees = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Fees"

    ws.append([
        "ID",
        "Student ID",
        "Amount",
        "Paid Date",
        "Status"
    ])

    for fee in fees:
        ws.append([
            fee["id"],
            fee["student_id"],
            fee["amount"],
            str(fee["paid_date"]),
            fee["status"]
        ])

    file_name = "fees_report.xlsx"
    wb.save(file_name)

    return FileResponse(file_name, filename=file_name)

@app.get("/export_teachers_excel")
def export_teachers_excel():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM teachers")
    teachers = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"

    ws.append([
        "ID",
        "Teacher Name",
        "Mobile",
        "Subject",
        "Salary"
    ])

    for teacher in teachers:
        ws.append([
            teacher["id"],
            teacher["teacher_name"],
            teacher["mobile"],
            teacher["subject"],
            teacher["salary"]
        ])

    file_name = "teachers_report.xlsx"
    wb.save(file_name)

    return FileResponse(file_name, filename=file_name)


@app.get("/export_attendance_excel")
def export_attendance_excel():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM attendance")
    attendance = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.append([
        "ID",
        "Student ID",
        "Date",
        "Status"
    ])

    for row in attendance:
        ws.append([
            row["id"],
            row["student_id"],
            row["date"],
            row["status"]
        ])

    file_name = "attendance_report.xlsx"
    wb.save(file_name)

    return FileResponse(file_name, filename=file_name)

@app.get("/export_results_excel")
def export_results_excel():
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM results")
    results = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    ws.append([
        "ID",
        "Student ID",
        "Subject",
        "Marks",
        "Total Marks",
        "Percentage",
        "Result"
    ])

    for result in results:
        ws.append([
            result["id"],
            result["student_id"],
            result["subject"],
            result["marks"],
            result["total_marks"],
            result["percentage"],
            result["result_status"]
        ])

    file_name = "results_report.xlsx"
    wb.save(file_name)

    return FileResponse(file_name, filename=file_name)

@app.get("/attendance_report", response_class=HTMLResponse)
def attendance_report(request: Request):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT s.student_name,
               a.date,
               a.status
        FROM attendance a
        JOIN students s
        ON a.student_id = s.id
        ORDER BY a.date DESC
    """)

    records = cursor.fetchall()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="records.html",
        context={
            "records": records
        }
    )

@app.get("/logout")
def logout(request: Request):
    request.session.clear()

    return RedirectResponse(
        url="/",
        status_code=303
    )

@app.get("/parent_dashboard/{student_id}", response_class=HTMLResponse)
def parent_dashboard(
    request: Request,
    student_id: int
):
    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(
        "SELECT * FROM students WHERE id=%s",
        (student_id,)
    )
    student = cursor.fetchone()

    cursor.execute("""
        SELECT *
        FROM fees
        WHERE student_id=%s
        ORDER BY id DESC
        LIMIT 1
    """, (student_id,))
    fee = cursor.fetchone()

    cursor.execute("""
        SELECT *
        FROM attendance
        WHERE student_id=%s
        ORDER BY date DESC
        LIMIT 5
    """, (student_id,))
    attendance = cursor.fetchall()

    cursor.execute("""
        SELECT *
        FROM results
        WHERE student_id=%s
    """, (student_id,))
    results = cursor.fetchall()

    conn.close()

    if not student:
        return {"error": "Student not found"}

    if not fee:
        fee = {
            "total_fee": 0,
            "paid_fee": 0,
            "pending_fee": 0
        }

    return templates.TemplateResponse(
        request=request,
        name="parent_dashboard.html",
        context={
            "student": student,
            "fee": fee,
            "attendance": attendance,
            "results": results
        }
    )

@app.get("/teacher_dashboard", response_class=HTMLResponse)
def teacher_dashboard(request: Request):

    if "user" not in request.session:
        return RedirectResponse("/", status_code=303)

    if request.session["role"] != "teacher":
        return RedirectResponse("/", status_code=303)

    conn = connect_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM students")
    students = cursor.fetchall()

    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="teacher_dashboard.html",
        context={
            "students": students
        }
    )   